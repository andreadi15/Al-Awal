# =======================
# FILE: services/excel_template_exporter.py
# =======================
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from copy import copy
import re

class ExcelTemplateExporter:
    """
    Export data ke Excel menggunakan sistem template + placeholder.
    
    Fitur:
    - Auto-detect placeholder row ({{key}})
    - Copy styling lengkap (font, alignment, border, fill)
    - Insert multiple rows dengan konsistensi styling
    - Safe border handling untuk semua sisi cell
    """
    
    def __init__(self, template_path):
        """
        Initialize exporter dengan template Excel
        
        Args:
            template_path (str): Path ke file template Excel (.xlsx)
        """
        self.template_path = template_path
        self.wb = None
        self.ws = None
        self.template_row_index = None
        self.placeholder_pattern = re.compile(r'\{\{(\w+)\}\}')
    
    def export(self, data_list, output_path, sheet_name=None):
        """
        Export data ke Excel menggunakan template
        
        Args:
            data_list (list): List of dict dengan key sesuai placeholder
            output_path (str): Path output file Excel
            sheet_name (str, optional): Nama sheet yang akan digunakan
            
        Returns:
            bool: True jika berhasil, False jika gagal
        """
        try:
            # Load template
            self.wb = load_workbook(self.template_path)
            self.ws = self.wb.active if not sheet_name else self.wb[sheet_name]
            
            # Cari template row (row dengan placeholder)
            self.template_row_index = self._find_template_row()
            
            if not self.template_row_index:
                print("[ERROR] Template row tidak ditemukan! Pastikan ada placeholder {{key}}")
                return False
            
            print(f"[INFO] Template row ditemukan di baris: {self.template_row_index}")
            
            # 🔥 Simpan styling template row DAN nilai placeholder
            template_styles = self._save_row_styles(self.template_row_index)
            template_values = self._save_row_values(self.template_row_index)
            
            # 🔥 LOGIC BARU: Fill data pertama ke template row
            self._fill_row_data(self.template_row_index, data_list[0])
            
            # 🔥 Insert row baru untuk data ke-2 dst (SETELAH template row)
            for i in range(1, len(data_list)):
                # Insert row baru SETELAH row terakhir yang sudah diisi
                insert_position = self.template_row_index + i
                self.ws.insert_rows(insert_position)
                
                # Copy styling ke row baru
                self._apply_row_styles(insert_position, template_styles)
                
                # Restore placeholder dulu (untuk konsistensi)
                self._restore_row_values(insert_position, template_values)
                
                # Fill data ke row baru
                self._fill_row_data(insert_position, data_list[i])
            
            # Save workbook
            self.wb.save(output_path)
            print(f"[SUCCESS] File berhasil disimpan: {output_path}")
            print(f"[INFO] Total data: {len(data_list)} baris")
            return True
            
        except Exception as e:
            print(f"[ERROR] Export gagal: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
        finally:
            if self.wb:
                self.wb.close()
    
    def _find_template_row(self):
        """
        Cari row yang mengandung placeholder {{key}}
        
        Returns:
            int: Index row template (1-based), atau None jika tidak ditemukan
        """
        for row_idx, row in enumerate(self.ws.iter_rows(), start=1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if self.placeholder_pattern.search(cell.value):
                        return row_idx
        return None
    
    def _save_row_styles(self, row_index):
        """
        Simpan semua styling dari row template
        
        Args:
            row_index (int): Index row (1-based)
            
        Returns:
            dict: Dictionary berisi styling per kolom
        """
        styles = {}
        for cell in self.ws[row_index]:
            col_idx = cell.column
            
            # Copy style dengan deep copy untuk menghindari shared reference
            styles[col_idx] = {
                'font': self._copy_font(cell.font),
                'border': self._copy_border(cell.border),
                'fill': self._copy_fill(cell.fill),
                'alignment': self._copy_alignment(cell.alignment),
                'number_format': cell.number_format
            }
        
        return styles
    
    def _save_row_values(self, row_index):
        """
        Simpan nilai placeholder dari row template
        
        Args:
            row_index (int): Index row (1-based)
            
        Returns:
            dict: Dictionary berisi nilai per kolom
        """
        values = {}
        for cell in self.ws[row_index]:
            col_idx = cell.column
            values[col_idx] = cell.value
        return values
    
    def _restore_row_values(self, row_index, values):
        """
        Restore nilai placeholder ke row baru
        
        Args:
            row_index (int): Index row (1-based)
            values (dict): Dictionary nilai dari _save_row_values()
        """
        for cell in self.ws[row_index]:
            col_idx = cell.column
            if col_idx in values:
                cell.value = values[col_idx]
    
    def _apply_row_styles(self, row_index, styles):
        """
        Apply styling ke row berdasarkan template
        
        Args:
            row_index (int): Index row target (1-based)
            styles (dict): Dictionary styling dari _save_row_styles()
        """
        for cell in self.ws[row_index]:
            col_idx = cell.column
            
            if col_idx in styles:
                style = styles[col_idx]
                cell.font = style['font']
                cell.border = style['border']
                cell.fill = style['fill']
                cell.alignment = style['alignment']
                cell.number_format = style['number_format']
    
    def _fill_row_data(self, row_index, data):
        """
        Isi data ke row dengan replace placeholder
        
        Args:
            row_index (int): Index row (1-based)
            data (dict): Data yang akan diisi
        """
        for cell in self.ws[row_index]:
            if cell.value and isinstance(cell.value, str):
                # Replace semua placeholder di cell
                new_value = cell.value
                
                for match in self.placeholder_pattern.finditer(cell.value):
                    placeholder = match.group(0)  # {{key}}
                    key = match.group(1)          # key
                    
                    # Get value dari data, default ke string kosong
                    value = data.get(key, '')
                    new_value = new_value.replace(placeholder, str(value))
                
                cell.value = new_value
    
    # ==================== STYLE COPY HELPERS ====================
    
    @staticmethod
    def _copy_font(font):
        """Deep copy Font object"""
        if not font:
            return Font()
        return Font(
            name=font.name,
            size=font.size,
            bold=font.bold,
            italic=font.italic,
            vertAlign=font.vertAlign,
            underline=font.underline,
            strike=font.strike,
            color=copy(font.color)
        )
    
    @staticmethod
    def _copy_border(border):
        """Deep copy Border object dengan semua sisi"""
        if not border:
            return Border()
        
        return Border(
            left=ExcelTemplateExporter._copy_side(border.left),
            right=ExcelTemplateExporter._copy_side(border.right),
            top=ExcelTemplateExporter._copy_side(border.top),
            bottom=ExcelTemplateExporter._copy_side(border.bottom),
            diagonal=ExcelTemplateExporter._copy_side(border.diagonal),
            diagonal_direction=border.diagonal_direction,
            outline=border.outline,
            vertical=ExcelTemplateExporter._copy_side(border.vertical),
            horizontal=ExcelTemplateExporter._copy_side(border.horizontal)
        )
    
    @staticmethod
    def _copy_side(side):
        """Deep copy Side object (untuk border)"""
        if not side:
            return Side()
        return Side(
            style=side.style,
            color=copy(side.color)
        )
    
    @staticmethod
    def _copy_fill(fill):
        """Deep copy PatternFill object"""
        if not fill or fill.fill_type is None:
            return PatternFill()
        return PatternFill(
            fill_type=fill.fill_type,
            start_color=copy(fill.start_color),
            end_color=copy(fill.end_color)
        )
    
    @staticmethod
    def _copy_alignment(alignment):
        """Deep copy Alignment object"""
        if not alignment:
            return Alignment()
        return Alignment(
            horizontal=alignment.horizontal,
            vertical=alignment.vertical,
            text_rotation=alignment.text_rotation,
            wrap_text=alignment.wrap_text,
            shrink_to_fit=alignment.shrink_to_fit,
            indent=alignment.indent
        )