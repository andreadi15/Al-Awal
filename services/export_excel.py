from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from config import DEBUG
from copy import copy
import re, logging

class export_Excel:
    def __init__(self, template_path):
        self.template_path = template_path
        self.wb = None
        self.ws = None
        self.template_row_index = None
        self.placeholder_pattern = re.compile(r'\{\{(\w+)\}\}')
    
    def export(self, data_list, output_path, sheet_name=None):
        try:
            self.wb = load_workbook(self.template_path)
            self.ws = self.wb.active if not sheet_name else self.wb[sheet_name]
            
            self.template_row_index = self._find_template_row()
            
            if not self.template_row_index:
                if DEBUG:
                    logging.error("[ERROR] Template row tidak ditemukan! Pastikan ada placeholder {{key}}")
                return False
            if DEBUG:
                logging.info(f"[INFO] Template row ditemukan di baris: {self.template_row_index}")
            
            template_styles = self._save_row_styles(self.template_row_index)
            template_values = self._save_row_values(self.template_row_index)
            
            self._fill_row_data(self.template_row_index, data_list[0])
            
            for i in range(1, len(data_list)):
                insert_position = self.template_row_index + i
                self.ws.insert_rows(insert_position)
                self._apply_row_styles(insert_position, template_styles)
                self._restore_row_values(insert_position, template_values)
                self._fill_row_data(insert_position, data_list[i])
            
            self.wb.save(output_path)
            
            import win32com.client as win32

            excel = win32.DispatchEx("Excel.Application")  
            excel.Visible = False

            wb = excel.Workbooks.Open(output_path)
            ws = wb.ActiveSheet
            ws.Rows.AutoFit()

            wb.Save()
            wb.Close(False)
            excel.Quit()
            if DEBUG:
                logging.info(f"[SUCCESS] File berhasil disimpan: {output_path}")
                logging.info(f"[INFO] Total data: {len(data_list)} baris")
            return True
            
        except Exception as e:
            if DEBUG:
                logging.error(f"[ERROR] Export gagal: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
        finally:
            if self.wb:
                self.wb.close()
    
    def _find_template_row(self):
        for row_idx, row in enumerate(self.ws.iter_rows(), start=1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if self.placeholder_pattern.search(cell.value):
                        return row_idx
        return None
    
    def _save_row_styles(self, row_index):
        styles = {}
        for cell in self.ws[row_index]:
            col_idx = cell.column
            
            styles[col_idx] = {
                'font': self._copy_font(cell.font),
                'border': self._copy_border(cell.border),
                'fill': self._copy_fill(cell.fill),
                'alignment': self._copy_alignment(cell.alignment),
                'number_format': cell.number_format
            }
        
        return styles
    
    def _save_row_values(self, row_index):
        values = {}
        for cell in self.ws[row_index]:
            col_idx = cell.column
            values[col_idx] = cell.value
        return values
    
    def _restore_row_values(self, row_index, values):
        for cell in self.ws[row_index]:
            col_idx = cell.column
            if col_idx in values:
                cell.value = values[col_idx]
    
    def _apply_row_styles(self, row_index, styles):
        for cell in self.ws[row_index]:
            col_idx = cell.column
            
            if col_idx in styles:
                style = styles[col_idx]
                cell.font = style['font']
                cell.border = style['border']
                cell.fill = style['fill']
                cell.alignment = style['alignment']
                cell.number_format = style['number_format']
    
    def _fill_row_data(self, row_index, data):
        for cell in self.ws[row_index]:
            if cell.value and isinstance(cell.value, str):
                new_value = cell.value
                
                for match in self.placeholder_pattern.finditer(cell.value):
                    placeholder = match.group(0)  
                    key = match.group(1)         
                    
                    value = data.get(key, '')
                    new_value = new_value.replace(placeholder, str(value))
                
                cell.value = new_value
    
    
    
    # ==================== STYLE COPY HELPERS ====================
    @staticmethod
    def _copy_font(font):
        if not font:
            return Font()
        return Font(
            name=font.name,
            size=font.size,
            bold=font.bold,
            italic=font.italic,
            vertAlign=font.vertAlign,
            underline=font.underline,
            strike=font.strike,
            color=copy(font.color)
        )
    
    @staticmethod
    def _copy_border(border):
        if not border:
            return Border()
        
        return Border(
            left=export_Excel._copy_side(border.left),
            right=export_Excel._copy_side(border.right),
            top=export_Excel._copy_side(border.top),
            bottom=export_Excel._copy_side(border.bottom),
            diagonal=export_Excel._copy_side(border.diagonal),
            diagonal_direction=border.diagonal_direction,
            outline=border.outline,
            vertical=export_Excel._copy_side(border.vertical),
            horizontal=export_Excel._copy_side(border.horizontal)
        )
    
    @staticmethod
    def _copy_side(side):
        if not side:
            return Side()
        return Side(
            style=side.style,
            color=copy(side.color)
        )
    
    @staticmethod
    def _copy_fill(fill):
        if not fill or fill.fill_type is None:
            return PatternFill()
        return PatternFill(
            fill_type=fill.fill_type,
            start_color=copy(fill.start_color),
            end_color=copy(fill.end_color)
        )
    
    @staticmethod
    def _copy_alignment(alignment):
        if not alignment:
            return Alignment()
        return Alignment(
            horizontal=alignment.horizontal,
            vertical=alignment.vertical,
            text_rotation=alignment.text_rotation,
            wrap_text=alignment.wrap_text,
            shrink_to_fit=alignment.shrink_to_fit,
            indent=alignment.indent
        )